"""Read experimental data into a ``Table`` that remembers cell provenance.

Every value keeps the coordinate it came from, such as ``Sheet1!C4``, so the
fidelity verifier can prove that a plotted number really appears in the source
file rather than in a template's demo data.

Supported inputs: ``.xlsx``/``.xlsm`` (openpyxl), ``.xls`` (xlrd), ``.csv``,
``.tsv``, ``.txt``, ``.json``, ``.jsonl``/``.ndjson``, ``.md``/``.markdown``.
"""

from __future__ import annotations

import csv
import datetime as dt
import json
import math
import re
from dataclasses import dataclass, field
from enum import StrEnum
from pathlib import Path
from typing import Any, Final, Iterable, Iterator, Sequence

import numpy as np
from numpy.typing import NDArray

SPREADSHEET_SUFFIXES: Final[frozenset[str]] = frozenset({".xlsx", ".xlsm"})
LEGACY_SPREADSHEET_SUFFIXES: Final[frozenset[str]] = frozenset({".xls"})
DELIMITED_SUFFIXES: Final[frozenset[str]] = frozenset({".csv", ".tsv", ".txt", ".tab"})
JSON_SUFFIXES: Final[frozenset[str]] = frozenset({".json"})
JSONL_SUFFIXES: Final[frozenset[str]] = frozenset({".jsonl", ".ndjson"})
MARKDOWN_SUFFIXES: Final[frozenset[str]] = frozenset({".md", ".markdown"})

SUPPORTED_SUFFIXES: Final[frozenset[str]] = (
    SPREADSHEET_SUFFIXES
    | LEGACY_SPREADSHEET_SUFFIXES
    | DELIMITED_SUFFIXES
    | JSON_SUFFIXES
    | JSONL_SUFFIXES
    | MARKDOWN_SUFFIXES
)

_TRUE_WORDS: Final[frozenset[str]] = frozenset(
    {"true", "yes", "y", "t", "1", "✓", "✔", "x", "present", "有"}
)
_FALSE_WORDS: Final[frozenset[str]] = frozenset(
    {"false", "no", "n", "f", "0", "×", "✗", "-", "--", "absent", "无"}
)
_MISSING_WORDS: Final[frozenset[str]] = frozenset(
    {"", "na", "n/a", "nan", "none", "null", "nd", "n.d.", "--", "—", "-"}
)

# 0.45 +/- 0.02, 0.45 (0.02), 0.45 ± 0.02
_UNCERTAINTY = re.compile(
    r"^(?P<value>[-+]?[\d.,]+(?:[eE][-+]?\d+)?)\s*"
    r"(?:±|\+/-|\+-|\\pm)\s*(?P<error>[\d.,]+(?:[eE][-+]?\d+)?)$"
)
_PARENTHESISED = re.compile(
    r"^(?P<value>[-+]?[\d.,]+(?:[eE][-+]?\d+)?)\s*"
    r"\(\s*(?P<error>[\d.,]+(?:[eE][-+]?\d+)?)\s*\)$"
)
_NUMBER = re.compile(r"^[-+]?(?:\d+\.?\d*|\.\d+)(?:[eE][-+]?\d+)?$")
_EMPHASIS = re.compile(r"^(?:\*\*|__|<b>)(?P<inner>.+?)(?:\*\*|__|</b>)$")
_TRAILING_STARS = re.compile(r"[*\u2020\u2021#†‡]+$")
_DATE_HINT = re.compile(
    r"^\d{4}[-/]\d{1,2}(?:[-/]\d{1,2})?$|^\d{1,2}[-/]\d{1,2}[-/]\d{2,4}$"
)
_MONTH_HINT = re.compile(
    r"\b(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)",
    re.IGNORECASE,
)


class ColumnKind(StrEnum):
    """What a column holds, after inspecting its values."""

    NUMERIC = "numeric"
    CATEGORICAL = "categorical"
    BOOLEAN = "boolean"
    TEMPORAL = "temporal"
    EMPTY = "empty"


@dataclass(frozen=True, slots=True)
class Measurement:
    """A parsed numeric cell, with anything extra the text carried."""

    value: float
    error: float | None = None
    is_percent: bool = False
    emphasised: bool = False
    """True when the source marked this value in bold, as benchmark tables do
    for the winning method."""


def column_letter(index: int) -> str:
    """A1-style column letter for a zero-based column index."""

    if index < 0:
        raise ValueError("column index must be non-negative")
    letters = ""
    index += 1
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(ord("A") + remainder) + letters
    return letters


def _clean(text: str) -> str:
    return text.replace("\u00a0", " ").strip()


def parse_measurement(raw: Any) -> Measurement | None:
    """Parse a cell into a :class:`Measurement`, or ``None`` if not numeric.

    Understands plain numbers, thousands separators, percentages, ``mean ± sd``
    and ``mean (sd)`` pairs, markdown bold emphasis, and trailing significance
    markers such as ``0.42**``.
    """

    if raw is None:
        return None
    if isinstance(raw, bool):
        return None
    if isinstance(raw, (int, float)):
        value = float(raw)
        return None if math.isnan(value) else Measurement(value=value)
    if isinstance(raw, (dt.datetime, dt.date, dt.time)):
        return None

    text = _clean(str(raw))
    if text.lower() in _MISSING_WORDS:
        return None

    emphasised = False
    match = _EMPHASIS.match(text)
    if match:
        emphasised = True
        text = _clean(match.group("inner"))

    is_percent = text.endswith("%")
    if is_percent:
        text = _clean(text[:-1])

    text = _TRAILING_STARS.sub("", text).strip()
    if not text:
        return None

    for pattern in (_UNCERTAINTY, _PARENTHESISED):
        found = pattern.match(text)
        if found:
            value = _to_float(found.group("value"))
            error = _to_float(found.group("error"))
            if value is None or error is None:
                continue
            return Measurement(
                value=value,
                error=abs(error),
                is_percent=is_percent,
                emphasised=emphasised,
            )

    value = _to_float(text)
    if value is None:
        return None
    return Measurement(value=value, is_percent=is_percent, emphasised=emphasised)


def _to_float(text: str) -> float | None:
    candidate = _clean(text).replace(",", "")
    if not _NUMBER.match(candidate):
        return None
    try:
        value = float(candidate)
    except ValueError:
        return None
    return None if math.isnan(value) else value


def parse_boolean(raw: Any) -> bool | None:
    """Parse tick/cross, yes/no and 0/1 style membership flags."""

    if raw is None:
        return None
    if isinstance(raw, bool):
        return raw
    text = _clean(str(raw)).lower()
    if text in _MISSING_WORDS - {"-", "--"}:
        return None
    if text in _TRUE_WORDS:
        return True
    if text in _FALSE_WORDS:
        return False
    return None


def is_missing(raw: Any) -> bool:
    if raw is None:
        return True
    if isinstance(raw, str):
        return _clean(raw).lower() in _MISSING_WORDS
    if isinstance(raw, float):
        return math.isnan(raw)
    return False


def _looks_temporal(raw: Any) -> bool:
    if isinstance(raw, (dt.datetime, dt.date)):
        return True
    if not isinstance(raw, str):
        return False
    text = _clean(raw)
    return bool(_DATE_HINT.match(text) or _MONTH_HINT.search(text))


@dataclass(frozen=True, slots=True)
class Column:
    """One column of a table, with a source reference per row."""

    name: str
    values: tuple[Any, ...]
    references: tuple[str, ...]
    header_reference: str = ""

    def __post_init__(self) -> None:
        if len(self.values) != len(self.references):
            raise ValueError(
                f"column {self.name!r}: {len(self.values)} values but "
                f"{len(self.references)} references"
            )

    def __len__(self) -> int:
        return len(self.values)

    @property
    def kind(self) -> ColumnKind:
        present = [value for value in self.values if not is_missing(value)]
        if not present:
            return ColumnKind.EMPTY
        booleans = sum(1 for value in present if parse_boolean(value) is not None)
        numerics = sum(1 for value in present if parse_measurement(value) is not None)
        temporals = sum(1 for value in present if _looks_temporal(value))
        total = len(present)
        # Booleans first: 0/1 parses as numeric too, but a two-valued 0/1 column
        # is a membership flag, not a measurement.
        if booleans == total and len(self.distinct()) <= 2:
            return ColumnKind.BOOLEAN
        if temporals >= 0.8 * total:
            return ColumnKind.TEMPORAL
        if numerics >= 0.8 * total:
            return ColumnKind.NUMERIC
        return ColumnKind.CATEGORICAL

    def distinct(self) -> tuple[Any, ...]:
        seen: list[Any] = []
        for value in self.values:
            if is_missing(value):
                continue
            key = _clean(str(value))
            if key not in {_clean(str(item)) for item in seen}:
                seen.append(value)
        return tuple(seen)

    def measurements(self) -> tuple[Measurement | None, ...]:
        return tuple(parse_measurement(value) for value in self.values)

    def numeric(self) -> NDArray[np.float64]:
        """Values as floats, with NaN where a cell is missing or non-numeric."""

        parsed = self.measurements()
        return np.asarray(
            [float("nan") if item is None else item.value for item in parsed],
            dtype=float,
        )

    def errors(self) -> NDArray[np.float64] | None:
        """Parsed uncertainties, or ``None`` when the column carries none."""

        parsed = self.measurements()
        if not any(item is not None and item.error is not None for item in parsed):
            return None
        return np.asarray(
            [
                float("nan") if item is None or item.error is None else item.error
                for item in parsed
            ],
            dtype=float,
        )

    def labels(self) -> tuple[str, ...]:
        return tuple("" if is_missing(value) else _clean(str(value)) for value in self.values)

    def emphasised_rows(self) -> tuple[int, ...]:
        return tuple(
            index
            for index, item in enumerate(self.measurements())
            if item is not None and item.emphasised
        )

    def percent_like(self) -> bool:
        parsed = [item for item in self.measurements() if item is not None]
        return bool(parsed) and all(item.is_percent for item in parsed)


@dataclass(frozen=True, slots=True)
class Table:
    """A rectangular block of data plus where it came from."""

    name: str
    source: Path
    columns: tuple[Column, ...]
    notes: tuple[str, ...] = field(default=())

    def __post_init__(self) -> None:
        if not self.columns:
            raise ValueError(f"table {self.name!r} has no columns")
        lengths = {len(column) for column in self.columns}
        if len(lengths) != 1:
            raise ValueError(f"table {self.name!r} has ragged columns: {sorted(lengths)}")

    @property
    def n_rows(self) -> int:
        return len(self.columns[0])

    @property
    def n_columns(self) -> int:
        return len(self.columns)

    @property
    def column_names(self) -> tuple[str, ...]:
        return tuple(column.name for column in self.columns)

    def column(self, name: str) -> Column:
        for column in self.columns:
            if column.name == name:
                return column
        raise KeyError(
            f"table {self.name!r} has no column {name!r}; "
            f"available: {', '.join(self.column_names)}"
        )

    def columns_of(self, kind: ColumnKind) -> tuple[Column, ...]:
        return tuple(column for column in self.columns if column.kind == kind)

    def numeric_columns(self) -> tuple[Column, ...]:
        return self.columns_of(ColumnKind.NUMERIC)

    def label_columns(self) -> tuple[Column, ...]:
        return tuple(
            column
            for column in self.columns
            if column.kind in {ColumnKind.CATEGORICAL, ColumnKind.TEMPORAL}
        )

    def matrix(self, columns: Sequence[Column] | None = None) -> NDArray[np.float64]:
        """Numeric columns as a ``(rows, columns)`` array."""

        chosen = tuple(columns) if columns is not None else self.numeric_columns()
        if not chosen:
            return np.zeros((self.n_rows, 0), dtype=float)
        return np.column_stack([column.numeric() for column in chosen])

    def cell_reference(self, column_name: str, row: int) -> str:
        return self.column(column_name).references[row]

    def describe(self) -> str:
        parts = [f"{self.name} ({self.n_rows} rows x {self.n_columns} columns)"]
        for column in self.columns:
            parts.append(f"  {column.name}: {column.kind}")
        return "\n".join(parts)


# ---------------------------------------------------------------- readers


def _header_and_body(
    grid: list[list[Any]],
) -> tuple[list[str], list[list[Any]], int]:
    """Split a raw grid into header labels and body rows.

    Returns the header labels, the body rows, and the zero-based grid row index
    the header was taken from so provenance references stay correct.
    """

    trimmed = [row for row in grid if any(not is_missing(cell) for cell in row)]
    if not trimmed:
        return [], [], 0

    header_index = 0
    for index, row in enumerate(grid):
        if any(not is_missing(cell) for cell in row):
            header_index = index
            break

    header_row = grid[header_index]
    numeric_in_header = sum(
        1
        for cell in header_row
        if not is_missing(cell) and parse_measurement(cell) is not None
    )
    filled_in_header = sum(1 for cell in header_row if not is_missing(cell))
    # A header row is mostly text; a grid that starts with numbers has none.
    has_header = filled_in_header > 0 and numeric_in_header <= 0.5 * filled_in_header

    width = max(len(row) for row in grid)
    if has_header:
        labels = [
            _clean(str(header_row[index])) if index < len(header_row) and not is_missing(header_row[index]) else ""
            for index in range(width)
        ]
        body_start = header_index + 1
    else:
        labels = ["" for _ in range(width)]
        body_start = header_index

    labels = _deduplicate(
        [label or f"column_{column_letter(index)}" for index, label in enumerate(labels)]
    )
    body = [row + [None] * (width - len(row)) for row in grid[body_start:]]
    body = [row for row in body if any(not is_missing(cell) for cell in row)]
    return labels, body, header_index


def _deduplicate(labels: Sequence[str]) -> list[str]:
    seen: dict[str, int] = {}
    result: list[str] = []
    for label in labels:
        if label in seen:
            seen[label] += 1
            result.append(f"{label}_{seen[label]}")
        else:
            seen[label] = 0
            result.append(label)
    return result


def _table_from_grid(
    name: str, source: Path, grid: list[list[Any]], *, row_offset: int = 0
) -> Table | None:
    labels, body, header_index = _header_and_body(grid)
    if not labels or not body:
        return None
    columns: list[Column] = []
    header_row_number = row_offset + header_index + 1
    first_body_row = row_offset + header_index + 2
    for index, label in enumerate(labels):
        values = tuple(row[index] if index < len(row) else None for row in body)
        references = tuple(
            f"{name}!{column_letter(index)}{first_body_row + offset}"
            for offset in range(len(values))
        )
        columns.append(
            Column(
                name=label,
                values=values,
                references=references,
                header_reference=f"{name}!{column_letter(index)}{header_row_number}",
            )
        )
    # Drop columns that are entirely empty; they are usually spacer columns.
    columns = [column for column in columns if column.kind != ColumnKind.EMPTY]
    if not columns:
        return None
    return Table(name=name, source=source, columns=tuple(columns))


def _read_xlsx(path: Path) -> list[Table]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency is declared
        raise ImportError(
            "reading .xlsx requires openpyxl; install it with "
            "`uv pip install openpyxl`"
        ) from exc

    workbook = load_workbook(path, data_only=True, read_only=True)
    tables: list[Table] = []
    try:
        for sheet in workbook.worksheets:
            grid = [list(row) for row in sheet.iter_rows(values_only=True)]
            table = _table_from_grid(sheet.title, path, grid)
            if table is not None:
                tables.append(table)
    finally:
        workbook.close()
    return tables


def _read_xls(path: Path) -> list[Table]:
    try:
        import xlrd
    except ImportError as exc:
        raise ImportError(
            "reading legacy .xls requires xlrd; install it with "
            "`uv pip install xlrd`, or re-save the file as .xlsx"
        ) from exc

    book = xlrd.open_workbook(path)
    tables: list[Table] = []
    for sheet in book.sheets():
        grid = [
            [sheet.cell_value(row, col) for col in range(sheet.ncols)]
            for row in range(sheet.nrows)
        ]
        table = _table_from_grid(sheet.name, path, grid)
        if table is not None:
            tables.append(table)
    return tables


def _sniff_delimiter(sample: str, suffix: str) -> str:
    if suffix in {".tsv", ".tab"}:
        return "\t"
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except csv.Error:
        return "\t" if sample.count("\t") > sample.count(",") else ","


def _read_delimited(path: Path) -> list[Table]:
    text = path.read_text(encoding="utf-8-sig", errors="replace")
    if not text.strip():
        return []
    if path.suffix.lower() in MARKDOWN_SUFFIXES or "|" in text.splitlines()[0]:
        markdown = _read_markdown_text(text, path)
        if markdown:
            return markdown
    delimiter = _sniff_delimiter(text[:8192], path.suffix.lower())
    grid = [list(row) for row in csv.reader(text.splitlines(), delimiter=delimiter)]
    table = _table_from_grid(path.stem, path, grid)
    return [table] if table is not None else []


def _split_markdown_row(line: str) -> list[str]:
    stripped = line.strip()
    if stripped.startswith("|"):
        stripped = stripped[1:]
    if stripped.endswith("|"):
        stripped = stripped[:-1]
    return [_clean(cell) for cell in stripped.split("|")]


def _is_markdown_separator(line: str) -> bool:
    cells = _split_markdown_row(line)
    if not cells:
        return False
    return all(re.fullmatch(r":?-{2,}:?", cell.replace(" ", "")) for cell in cells)


def _read_markdown_text(text: str, path: Path) -> list[Table]:
    """Extract every pipe table in a markdown document, in order."""

    lines = text.splitlines()
    tables: list[Table] = []
    heading = ""
    index = 0
    while index < len(lines):
        line = lines[index]
        if line.lstrip().startswith("#"):
            heading = _clean(line.lstrip("#"))
            index += 1
            continue
        if (
            "|" in line
            and index + 1 < len(lines)
            and _is_markdown_separator(lines[index + 1])
        ):
            header = _split_markdown_row(line)
            body: list[list[Any]] = []
            cursor = index + 2
            while cursor < len(lines) and "|" in lines[cursor]:
                row = _split_markdown_row(lines[cursor])
                if row and any(cell for cell in row):
                    body.append(row + [None] * (len(header) - len(row)))
                cursor += 1
            name = heading or f"{path.stem}_table_{len(tables) + 1}"
            grid = [header, *body]
            table = _table_from_grid(name, path, grid)
            if table is not None:
                tables.append(table)
            index = cursor
            continue
        index += 1
    return tables


def _read_markdown(path: Path) -> list[Table]:
    return _read_markdown_text(
        path.read_text(encoding="utf-8-sig", errors="replace"), path
    )


def _flatten(prefix: str, node: Any) -> Iterator[tuple[str, Any]]:
    if isinstance(node, dict):
        for key, value in node.items():
            child = f"{prefix}.{key}" if prefix else str(key)
            if isinstance(value, (dict, list)):
                yield from _flatten(child, value)
            else:
                yield child, value
    elif isinstance(node, list):
        for index, value in enumerate(node):
            child = f"{prefix}[{index}]"
            if isinstance(value, (dict, list)):
                yield from _flatten(child, value)
            else:
                yield child, value
    else:
        yield prefix, node


def _records_to_table(
    name: str, source: Path, records: Sequence[dict[str, Any]], pointer: str
) -> Table | None:
    labels: list[str] = []
    for record in records:
        for key in record:
            if key not in labels:
                labels.append(key)
    if not labels:
        return None
    columns: list[Column] = []
    for label in labels:
        values = tuple(record.get(label) for record in records)
        references = tuple(
            f"{pointer}[{index}].{label}" for index in range(len(records))
        )
        columns.append(Column(name=label, values=values, references=references))
    columns = [column for column in columns if column.kind != ColumnKind.EMPTY]
    if not columns:
        return None
    return Table(name=name, source=source, columns=tuple(columns))


def _find_record_arrays(node: Any, pointer: str = "$") -> Iterator[tuple[str, list[dict[str, Any]]]]:
    """Yield every list-of-objects in a JSON document, outermost first."""

    if isinstance(node, list) and node and all(isinstance(item, dict) for item in node):
        yield pointer, list(node)
        return
    if isinstance(node, dict):
        for key, value in node.items():
            yield from _find_record_arrays(value, f"{pointer}.{key}")
    elif isinstance(node, list):
        for index, value in enumerate(node):
            yield from _find_record_arrays(value, f"{pointer}[{index}]")


def _read_json(path: Path) -> list[Table]:
    payload = json.loads(path.read_text(encoding="utf-8-sig"))
    tables: list[Table] = []
    for pointer, records in _find_record_arrays(payload):
        name = pointer if pointer != "$" else path.stem
        table = _records_to_table(name, path, records, pointer)
        if table is not None:
            tables.append(table)
    if tables:
        return tables

    # No record arrays: treat the flattened scalar leaves as a key/value table.
    pairs = [(key, value) for key, value in _flatten("", payload) if key]
    if not pairs:
        return []
    keys = Column(
        name="key",
        values=tuple(key for key, _ in pairs),
        references=tuple(f"$.{key}" for key, _ in pairs),
    )
    values = Column(
        name="value",
        values=tuple(value for _, value in pairs),
        references=tuple(f"$.{key}" for key, _ in pairs),
    )
    return [Table(name=path.stem, source=path, columns=(keys, values))]


def _read_jsonl(path: Path) -> list[Table]:
    records: list[dict[str, Any]] = []
    for line in path.read_text(encoding="utf-8-sig").splitlines():
        stripped = line.strip()
        if not stripped:
            continue
        payload = json.loads(stripped)
        if isinstance(payload, dict):
            records.append(payload)
    table = _records_to_table(path.stem, path, records, "$")
    return [table] if table is not None else []


def read_tables(path: str | Path) -> tuple[Table, ...]:
    """Read every table in ``path``.

    Workbooks yield one table per sheet and markdown documents one per pipe
    table, so a single file can offer several candidates.
    """

    resolved = Path(path)
    if not resolved.exists():
        raise FileNotFoundError(f"no such data file: {resolved}")
    suffix = resolved.suffix.lower()

    if suffix in SPREADSHEET_SUFFIXES:
        tables = _read_xlsx(resolved)
    elif suffix in LEGACY_SPREADSHEET_SUFFIXES:
        tables = _read_xls(resolved)
    elif suffix in JSON_SUFFIXES:
        tables = _read_json(resolved)
    elif suffix in JSONL_SUFFIXES:
        tables = _read_jsonl(resolved)
    elif suffix in MARKDOWN_SUFFIXES:
        tables = _read_markdown(resolved)
    elif suffix in DELIMITED_SUFFIXES:
        tables = _read_delimited(resolved)
    else:
        raise ValueError(
            f"unsupported data format {suffix!r}; supported: "
            f"{', '.join(sorted(SUPPORTED_SUFFIXES))}"
        )

    if not tables:
        raise ValueError(f"found no usable table in {resolved}")
    return tuple(tables)


def read_table(path: str | Path, *, name: str | None = None) -> Table:
    """Read one table, choosing the largest when the file holds several."""

    tables = read_tables(path)
    if name is not None:
        for table in tables:
            if table.name == name:
                return table
        raise KeyError(
            f"{path} has no table named {name!r}; "
            f"available: {', '.join(table.name for table in tables)}"
        )
    return max(tables, key=lambda table: (table.n_rows * table.n_columns, -len(table.name)))


def read_many(paths: Iterable[str | Path]) -> tuple[Table, ...]:
    """Read every table from every path."""

    collected: list[Table] = []
    for path in paths:
        collected.extend(read_tables(path))
    return tuple(collected)
